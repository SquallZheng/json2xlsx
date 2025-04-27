import argparse
import copy
import json
from openpyxl.workbook import Workbook

def appendHeader(header, col):
    if not col:
        return
    if col not in header:
        header.append(col)

def isPrimitiveList(json):
    if type(json) is not list:
        return False
    for elem in json:
        if type(elem) in [list, dict]:
            return False
    return True

def dfsDepth(depths, json):
    depth = 0
    if type(json) is dict:
        for key, val in json.items():
            depth = max(depth, dfsDepth(depths, val))
        if len(json.items()) > 0:
            depth += 1
    elif type(json) is list:
        for elem in json:
            depth = max(depth, dfsDepth(depths, elem))
        if len(json) > 1:
            depth += 1
    else:
        depth = 1
    depths[str(json)] = depth
    return depth

# json -> dict | list | string
# dict -> { key : json (, key : json)* }
# list -> [ json (, json)* ]
# key -> string
def dfsHeader(header, json, depths):
    if type(json) is dict:
        depthSortedChilds = sorted(json.items(), key=lambda e: depths[str(e[1])])
        for key, val in depthSortedChilds:
            if type(val) not in [list, dict] or isPrimitiveList(val):
                appendHeader(header, key)
            else:
                dfsHeader(header, val, depths)
    elif type(json) is list:
        depthSortedChilds = sorted(json, key=lambda e: depths[str(e)])
        for elem in depthSortedChilds:
            dfsHeader(header, elem, depths)

def dfsRow(rows, row, json, header, headerIndex, depths):
    rowCount = 0
    headerIndexPrev = headerIndex[0]
    if type(json) is dict:
        depthSortedChilds = sorted(json.items(), key=lambda e: depths[str(e[1])])
        for key, val in depthSortedChilds:
            rowCount += dfsRow(rows, row, val, header, headerIndex, depths)
    elif type(json) is list:
        depthSortedChilds = sorted(json, key=lambda e: depths[str(e)])
        for elem in depthSortedChilds:
            rowCount += dfsRow(rows, row, elem, header, headerIndex, depths)
    else:
        row.append(json)
        headerIndex[0] += 1
        if headerIndex[0] == len(header):
            rows.append(copy.copy(row))
            rowCount += 1
    if rowCount > 0:
        row[:] = row[: len(row) - (headerIndex[0] - headerIndexPrev)]
        headerIndex[0] = headerIndexPrev
    return rowCount

def jsonToExcel(json, outputFileName, mergeCells=True, dictKeyHeader=None):
    if dictKeyHeader:
        dictKeyHeaderJson = []
        for key, val in json.items():
            dictKeyHeaderJson.append({dictKeyHeader: key, '_': val})
        json = dictKeyHeaderJson

    header = []
    rows = []
    row = []
    depths = {}
    dfsDepth(depths, json)
    dfsHeader(header, json, depths)
    dfsRow(rows, row, json, header, [0], depths)

    wb = Workbook()
    ws = wb.active
    ws.title = outputFileName

    ws.append(header)
    for row in rows:
        ws.append(row)

    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.startswith(("http://", "https://")):
                cell.hyperlink = val
                cell.style = "Hyperlink"

    if mergeCells:
        mergeStart = 1
        for col in range(1, ws.max_column + 1):
            for row in range(2, ws.max_row + 2):
                if row == ws.max_row + 1 or ws.cell(row, col).value != ws.cell(row - 1, col).value:
                    if row - mergeStart > 1:
                        ws.merge_cells(start_row=mergeStart, start_column=col, end_row=row - 1, end_column=col)
                    mergeStart = row

    wb.save(outputFileName)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='json2xlsx commands'
    )
    parser.add_argument('jsonFile', type=str, help='Path of input json file')
    parser.add_argument('outputFileName',type=str, help='Path of output xlsx file')
    parser.add_argument('notMergeCell', action='store_true', help='False to prohibit merge continuous identical cell vertically')
    parser.add_argument('--dictKeyHeader', type=str, default=None, help='If input json is a dictionary, extract its key as content of specified header column')
    args = parser.parse_args()
    with open(args.jsonFile, 'r', encoding='utf-8') as f:
        jsonData = json.load(f)
        jsonToExcel(jsonData, args.outputFileName, not args.notMergeCell, args.dictKeyHeader)


